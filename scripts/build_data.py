# -*- coding: utf-8 -*-
"""파일럿_분석결과.xlsx (분석결과 + 컬럼메타데이터 2개 시트) → src/data/ssi.json
   40개 컬럼 전부를 손실 없이 옮기고, 컬럼메타데이터를 그대로 동봉한다.
   표준화 공식은 별도 검산 스크립트(verify_formula.py)로 xlsx와 완전 일치 확인됨.
"""
import openpyxl, json, os, re

SRC = '/mnt/user-data/uploads/Downloads/파일럿_분석결과.xlsx'
OUT = os.path.join(os.path.dirname(__file__), 'data', 'ssi.json')

wb = openpyxl.load_workbook(SRC, data_only=True)

# ── 1. 컬럼메타데이터 시트 ────────────────────────────────────────────────
mrows = list(wb['컬럼메타데이터'].iter_rows(values_only=True))
MKEY = ['name', 'sector', 'kind', 'desc', 'unit', 'how', 'note']
COLMETA = [dict(zip(MKEY, [(x or '').strip() if isinstance(x, str) else x for x in r]))
           for r in mrows[1:] if r[0]]
META_BY = {c['name']: c for c in COLMETA}

# ── 2. 분석결과 시트 ─────────────────────────────────────────────────────
rows = list(wb['분석결과'].iter_rows(values_only=True))
HDR = list(rows[0])
DATA = [dict(zip(HDR, r)) for r in rows[1:] if r[0]]

num = lambda x: None if x in (None, '') else round(float(x), 4)

# 부문 · 지표 정의 (원자료 컬럼에서 자동 추출 + 메타데이터의 방향/연도)
SECTOR_NAME = {'S1': '공간구조효율성', 'S8': '지역사회건강도'}
sectors = {}
for s in ('S1', 'S8'):
    inds = []
    for c in HDR:
        if not c.startswith(f'{s}_원자료_'):
            continue
        label = c[len(f'{s}_원자료_'):]
        m = META_BY.get(c, {})
        note = m.get('note') or ''
        dirn = '+' if ('+1' in note or '+' == note[:1]) else '-'
        if '방향 +1' in note or '방향(+1' in note:
            dirn = '+'
        elif '방향 -1' in note or '방향(-1' in note:
            dirn = '-'
        yr = re.search(r'(\d{4})년', m.get('desc') or '')
        inds.append({
            'key': label, 'label': label,
            'unit': m.get('unit') or '', 'dir': dirn,
            'year': yr.group(1) if yr else '',
            'desc': m.get('desc') or '', 'col': c,
        })
    sectors[s] = {'name': SECTOR_NAME[s], 'inds': inds}

METHODS = [
    {'key': 'minmax',   'col': 'MinMax',   'label': 'Min-Max',
     'camp': '간격보존형', 'short': 'MM',
     'formula': '(x − min) / (max − min) × 100',
     'range': '0 ~ 100 (양끝 고정)',
     'note': '최솟값 0, 최댓값 100으로 고정. 값 간격을 선형으로 보존.'},
    {'key': 'distance', 'col': 'Distance', 'label': '거리기반',
     'camp': '간격보존형', 'short': 'DI',
     'formula': 'x / 전국평균 × 100',
     'range': '상한 없음 (실측 7.6 ~ 241.9)',
     'note': '100 = 전국평균. LQ(입지지수) × 100과 수학적으로 동일.'},
    {'key': 'pctrank',  'col': 'PctRank',  'label': '백분위순위',
     'camp': '순위전용형', 'short': 'PR',
     'formula': '평균순위(1기준) / N × 100',
     'range': '0 ~ 100 (균등 분포)',
     'note': '등수만 반영하고 값 간격은 버림. 분포 밀집을 그대로 드러냄.'},
    {'key': 'logistic', 'col': 'Logistic', 'label': '로지스틱',
     'camp': '간격보존형', 'short': 'LG',
     'formula': '100 / (1 + exp(−z)),  z = (x − 평균) / 표준편차',
     'range': '0 ~ 100 (양끝 도달 안함, 실측 18.8 ~ 80.8)',
     'note': '평균 근처 변별력이 높고 극단값을 압축. 값이 중간에 몰림.'},
]

out_rows = []
for d in DATA:
    row = {'sido': d['시도'], 'name': d['시군구']}
    for s in ('S1', 'S8'):
        blk = {
            'ci':   {m['key']: num(d.get(f'{s}_CI_{m["col"]}')) for m in METHODS},
            'rank': {m['key']: num(d.get(f'{s}_순위_{m["col"]}')) for m in METHODS},
            'ssiRange': num(d.get(f'{s}_SSI_range')),
            'ssiStd':   num(d.get(f'{s}_SSI_std')),
            'ssiCamp':  num(d.get(f'{s}_SSI_camp')),
            'flag':     d.get(f'{s}_민감구분'),
            # 검산용 중복 컬럼 (메타데이터가 "투명성을 위해 유지"라 명시)
            'repMinmax':  num(d.get(f'{s}_MinMax대표순위')),
            'repPctrank': num(d.get(f'{s}_PctRank대표순위')),
            'raw': {i['label']: num(d.get(i['col'])) for i in sectors[s]['inds']},
        }
        if s == 'S1':
            blk['tradeoff'] = (d.get('S1_트레이드오프_참고') == 'Y')
        row[s] = blk
    out_rows.append(row)

payload = {
    'meta': {
        'n': len(out_rows),
        'source': '파일럿_분석결과.xlsx',
        'sheets': ['분석결과 (229행 × 40열)', '컬럼메타데이터 (40행 × 7열)'],
        'guide': '표준화 방법론 및 민감도 진단 지침서 v2',
        'reversal': '방향 −1 지표는 x′ = max + min − x 로 반전 후 표준화 (4개 방법 모두 xlsx와 완전 일치 검산됨)',
    },
    'columns': COLMETA,
    'sectors': sectors,
    'methods': METHODS,
    'rows': out_rows,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False, separators=(',', ':'))
print('rows', len(out_rows), 'columns', len(COLMETA), 'bytes', os.path.getsize(OUT))
for s in sectors:
    print(s, sectors[s]['name'], [(i['label'], i['dir'], i['unit'], i['year']) for i in sectors[s]['inds']])
